from time import sleep
import openpyxl as opx
import pyautogui
import pyperclip
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os
from PIL import Image
import locale
import pandas as pd

def formatar(num):
    # Set the locale to use the thousancoluna separator ancoluna decimal point
    # that are appropriate for your region
    locale.setlocale(locale.LC_ALL, 'pt_BR.UTF-8')
    # Format the number as a string with two decimal places ancoluna a thousancoluna separator
    formatado = locale.format_string('%.2f', num, grouping=True)
    return formatado

def enviar_msg():
    pyautogui.click(186, 142)  # clicar no campo de contato
    sleep(1)
    pyautogui.write(envio[0][gerar_arquivo][2])  # escreve numero whats
    sleep(1)
    pyautogui.press("enter") # confirma numero whats
    sleep(2)
    pyautogui.click(667, 992)  #clica clips anexar
    sleep(1)
    pyautogui.click(674, 652) #clica anexar documento
    sleep(2)
    pyautogui.click(490, 61) #clica seletor path
    sleep(1)
    pyautogui.write(arquivo_path) #escreve path
    sleep(2)
    pyautogui.press('enter') #confirma
    pyautogui.click(586, 521, button='right')  # botao direito para colar
    pyautogui.click(656, 617) #clica colar
    sleep(1)
    pyautogui.press('enter') #confirma
    sleep(10)
    pyautogui.click(1847, 968) #envia pdf
    sleep(2)
    pyautogui.click(500, 144)
    sleep(1)

def date(data_plan):
    global ano, mes, dia
    ano = data_plan.year
    mes = data_plan.strftime('%m')
    dia = data_plan.strftime('%d')
    return ano, mes, dia

def merge_duplicates(lst):
    # Create a dictionary to store the lists with the same element in the third position
    groups = {}

    # Iterate over the list and group the lists with the same element in the third position
    for element in lst:
        # Get the third element of each sublist
        item = element[2]
        if item in groups:
            groups[item].append(element)
        else:
            groups[item] = [element]

    # Merge the lists with the same element in the third position
    merged_lists = []
    for group in groups.values():
        if len(group) > 1:
            # Set merged_list equal to the group of sublists with duplicate elements
            merged_list = group
            merged_lists.append(merged_list)
        else:
            merged_lists.append(group)

    return merged_lists


os.startfile('WhatsApp Web path')  # abrir whats

planilha = opx.load_workbook('sheet path', data_only=True)
dados = planilha["sheet name"]

#logo
logo_path = 'path: company logo'
logo = Image.open(logo_path)
width, height = logo.size
ratio = width/height
imagem_largura = 959
imagem_altura = imagem_largura / ratio
#assinatura
assinatura_path = 'path: digital signature that will be use'
assinatura = Image.open(assinatura_path)
awidth, aheight = assinatura.size
aratio = awidth/aheight
ass_largura = 484 * 1.7
ass_altura = ass_largura / aratio

todos = []
socio = []
contato = []
#dimensões página
largura = 1900
altura = 2500
margem = 200

#data
dia, mes, ano = '', '', ''

max_col = dados.max_column

for linha in range(1, 250):
    if dados.cell(row = linha + 1, column = 2).value == None:
        break
    else:
        numero_recibo = dados.cell(row = linha + 1, column = 1).value
        nome = dados.cell(row = linha + 1, column = 2).value
        whats = dados.cell(row = linha + 1, column = 3).value
        data_pagamento = date(dados.cell(row = linha + 1, column = 5).value)
        data = f'Data de Pagamento: {dia}/{mes}/{ano}'
        socio.append(numero_recibo)
        socio.append(nome)
        socio.append(whats)
        todos.append(socio)
        socio.append(f'{ano}_{mes}_{dia}')
        contato.append(socio)
        #criando arquivo
        pdf_nome = f'No {str(numero_recibo)}' + '_' + f'{ano}_{mes}_{dia}' + '_' + nome + '.pdf'
        cnv = canvas.Canvas(pdf_nome)
        cnv.setPageSize((largura, altura))
        socio = []
        for coluna in range(6, max_col + 1, 2):
            if dados.cell(row=linha + 1, column=coluna).value == None:
                pass
            else:
                debito = dados.cell(row=1, column=coluna).value
                if dados.cell(row=linha + 1, column=coluna + 1).value != None:
                    debito += " " + str(dados.cell(row=linha + 1, column=coluna + 1).value)
                valor_debito = dados.cell(row=linha + 1, column=coluna).value
                total = dados.cell(row = linha + 1, column = 4).value
                socio.append(debito)
                socio.append(valor_debito)
                socio.append(total)
                todos.append(socio)
                socio = []
        #gerar dados pdf
        #config textos
        pdfmetrics.registerFont(TTFont('Arial', 'Arial.ttf'))
        #escrever largura x altura
        for e in range(0, len(todos) - 1):
            cnv.drawImage(logo_path, 436, altura - 272, imagem_largura, imagem_altura, mask='auto')
            cnv.setFont('Arial',40)
            cnv.drawCentredString(largura / 2, altura - 303, '--------company name-------')
            cnv.setFont('Arial',32)
            cnv.drawCentredString(largura / 2, altura - 356, 'Write EIN, CNPJ, UTR etc.')
            cnv.drawCentredString(largura / 2, altura - 410, 'Write the address')
            cnv.line(0, altura - 433, largura, altura - 433)
            cnv.drawString(margem, altura - 478, (f'Recibo nº: {numero_recibo}'))
            cnv.drawString(largura - 740, altura - 478, data)
            cnv.setFont('Arial',40)
            cnv.drawString(margem, altura - 555, (f'Nome: {nome}'))
            cnv.setFont('Arial',46)
            cnv.drawString(436, altura - 655, 'DESCRIÇÃO')
            cnv.drawString(1315, altura - 655, 'VALOR')
            cnv.setFont('Arial',40)
            cnv.drawString(271, (altura - 655 - (e * 64 + 64)), (f'{str(todos[e + 1][0])}'))
            valor_prefix = 'R$ '
            valor_value = f'{formatar(float(todos[e + 1][1]))}'
            valor_field_width = 10  # Adjust this value as needed
            valor_prefix_aligned = valor_prefix.ljust(valor_field_width - len(valor_value))
            valor_value_aligned = valor_value.rjust(valor_field_width)
            valor_text = valor_prefix_aligned + valor_value_aligned
            valor_y = altura - 655 - (e * 64 + 64)
            valor_text_width = cnv.stringWidth(valor_text)
            cnv.drawString(1505 - valor_text_width, valor_y, valor_text)
        if len(todos) == 1:
            distancia = 200
            cnv.drawString(271, (altura - 655 - (distancia)), f'Total {79 * "."}')
            valor_value = f'{formatar(float(todos[0][e]))}'
            valor_prefix_alignecoluna = valor_prefix.ljust(valor_field_width - len(valor_value))
            valor_value_alignecoluna = valor_value.rjust(valor_field_width)
            valor_text = valor_prefix_alignecoluna + valor_value_aligned
            valor_y = altura - 655 - (distancia)
            valor_text_width = cnv.stringWidth(valor_text)
            cnv.drawString(1505 - valor_text_width, valor_y, valor_text)
        else:
            distancia = e * 64 + 200
            cnv.drawString(271, (altura - 655 - (distancia)), f'Total {79 * "."}')
            valor_value = f'{formatar(float(todos[e + 1][2]))}'
            valor_prefix_aligned = valor_prefix.ljust(valor_field_width - len(valor_value))
            valor_value_aligned = valor_value.rjust(valor_field_width)
            valor_text = valor_prefix_aligned + valor_value_aligned
            valor_y = altura - 655 - (distancia)
            valor_text_width = cnv.stringWidth(valor_text)
            cnv.drawString(1505 - valor_text_width, valor_y, valor_text)
        cnv.drawImage(assinatura_path, largura / 2 - (ass_largura / 2), valor_y - 350, ass_largura, ass_altura, mask='auto')
        cnv.setFillGray(0.8)  # Set the fill color to medium gray
        cnv.setFont('Arial', 25)
        cnv.drawString(0, valor_y - 398, f'{91 * "-"} CORTAR NA LINHA TRACEJADA {91 * "-"}')
        cnv.showPage
        todos = []
    cnv.save()

sleep(12)
# Test the function
merged_lists = merge_duplicates(contato)
envio = []

#print(merged_lists)
arquivo_path = 'pdf file path'
arquivo_nome = ""
plan_t = pd.ExcelFile('xlsx path')
plan = pd.read_excel(plan_t, 'sheet name') #ler aba correta
for lista in range(0, len(merged_lists)):
    envio.append(merged_lists[lista])
    for gerar_arquivo in range(0, len(envio[0])):
        if len(envio[0]) == 1:
            arquivo_nome += f'No {envio[0][gerar_arquivo][0]}' + "_" + envio[0][gerar_arquivo][3] + '_' + envio[0][gerar_arquivo][1] + " "
        else:
            arquivo_nome += f'"No {envio[0][gerar_arquivo][0]}' + "_" + envio[0][gerar_arquivo][3] + '_' + envio[0][gerar_arquivo][1] + '"' + " "
    pyperclip.copy(arquivo_nome)
    enviar_msg()
    pyautogui.click(394, 120)  # apaga contato
    envio = []
    arquivo_nome = ""











